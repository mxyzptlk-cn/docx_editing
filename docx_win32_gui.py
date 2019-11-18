#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# Author: Mxyzptlk
# Date: 2019/11/16

# v1.2
# pyinstaller -D -w -i C:\Users\Mxyzptlk\Documents\PycharmProjects\docx_editing\logo.ico docx_win32_gui.py

import PySimpleGUI as sg

import os
import gc
import datetime

import win32com
from win32com.client import Dispatch

from openpyxl import load_workbook as lb


def days(str1, str2):
    date1 = datetime.datetime.strptime(str1, "%Y.%m.%d")
    date2 = datetime.datetime.strptime(str2, "%Y.%m.%d")
    num = round((date1 - date2).days / 7)
    return num


class RemoteWord:
    def __init__(self, filename=None):
        self.xlApp = win32com.client.Dispatch('Word.Application')
        self.xlApp.Visible = 0  # 后台运行，不显示
        self.xlApp.DisplayAlerts = 0  # 不警告
        if filename:
            self.filename = filename
            if os.path.exists(self.filename):
                self.doc = self.xlApp.Documents.Open(filename)
            else:
                self.doc = self.xlApp.Documents.Add()  # 创建新的文档
                self.doc.SaveAs(filename)
        else:
            self.doc = self.xlApp.Documents.Add()
            self.filename = ''
    
    def add_doc_end(self, string):
        # '在文档末尾添加内容'
        c_range = self.doc.Range()
        c_range.InsertAfter('\n' + string)
    
    def add_doc_start(self, string):
        # '在文档开头添加内容'
        c_range = self.doc.Range(0, 0)
        c_range.InsertBefore(string + '\n')
    
    def insert_doc(self, in_position, string):
        # '在文档in_position位置添加内容'
        c_range = self.doc.Range(0, in_position)
        if in_position == 0:
            c_range.InsertAfter(string)
        else:
            c_range.InsertAfter('\n' + string)
    
    def replace_doc(self, string, new_string):
        # '替换文字'
        self.xlApp.Selection.Find.ClearFormatting()
        self.xlApp.Selection.Find.Replacement.ClearFormatting()
        # (string--搜索文本,
        # True--区分大小写,
        # True--完全匹配的单词，并非单词中的部分（全字匹配）,
        # True--使用通配符,
        # True--同音,
        # True--查找单词的各种形式,
        # True--向文档尾部搜索,
        # 1,
        # True--带格式的文本,
        # new_string--替换文本,
        # 2--替换个数（全部替换）
        self.xlApp.Selection.Find.Execute(string, False, False, False, False, False, True, 1, True, new_string, 2)
    
    def replace_header(self, string, new_string):
        # '页眉文字'
        self.xlApp.ActiveDocument.Sections[0].Headers[0].Range.Find.ClearFormatting()
        self.xlApp.ActiveDocument.Sections[0].Headers[0].Range.Find.Replacement.ClearFormatting()
        self.xlApp.ActiveDocument.Sections[0].Headers[0].Range.Find.Execute(string, False, False, False, False, False,
                                                                            True, 1, True,
                                                                            new_string, 2)
    
    def replace_docs(self, string, new_string):
        # '采用通配符匹配替换'
        self.xlApp.Selection.Find.ClearFormatting()
        self.xlApp.Selection.Find.Replacement.ClearFormatting()
        self.xlApp.Selection.Find.Execute(string, False, False, True, False, False, False, 1, False, new_string, 2)
    
    # 编辑表格内容
    def modify_tab(self, tab_num, row_num, cell_num, text):  # 表格编号从1开始
        self.doc.Tables(tab_num).Rows(row_num).Cells(cell_num).Range.Text = text
    
    # 编辑（含合并单元格的）表格内容
    def modify_tab2(self, tab_num, row_num, cell_num, text):  # 表格编号从1开始
        self.doc.Tables(tab_num).Cell(row_num, cell_num).Range.Text = text
    
    def ins_table_row(self, tab_num, row_num=1):
        for i in range(row_num):
            self.doc.Tables[tab_num - 1].Rows.Add()  # 源代码表格编号从0开始，这里+1，统一修改表格的编号设置
    
    def save(self):
        # '保存文档'
        self.doc.Save()
    
    def save_as(self, filename):
        # '文档另存为'
        self.doc.SaveAs(filename)
    
    def close(self):
        # '保存文件、关闭文件'
        self.xlApp.Documents.Close()
        self.xlApp.Quit()


def read_from_xlsx(file):
    ret = {'变量1': None, '变量2': None, '变量3': None, '变量4': None, '变量5': None, '变量6': None, '变量7': None,
           '变量8': None, '变量9': None, '变量10': None, '变量11': None, '变量12': None, '变量13': None, '变量14': None,
           '变量15': None, '变量16': None, '变量17': None, '变量18': None, '变量19': None, '变量20': None, '变量21': None}
    ret2 = []
    
    wb = lb(file, data_only=True, keep_vba=False)
    ws = wb.active
    
    if ws['C1'].value and ws['C2'].value:
        for i in range(1, 22):  # 变量1-21部分
            if ws['C' + str(i)].value:
                ret['变量' + str(i)] = ws['C' + str(i)].value
            else:
                ret['变量' + str(i)] = ' '
        if ws['F2'].value and ws['G2'].value:
            for i in range(2, 102):  # 系统-子系统部分
                temp = []
                if ws['F' + str(i)].value and ws['G' + str(i)].value:
                    temp.append(ws['F' + str(i)].value)
                    temp.append(ws['G' + str(i)].value)
                    temp.append(ws['H' + str(i)].value) if ws['H' + str(i)].value else temp.append('')
                    ret2.append(temp)
            return ret, ret2
        else:
            return ret, None
    else:
        return None


def comm_task(doc, ret):
    for i in reversed(range(1, 23)):
        if i == 1:
            doc.replace_header('变量1', ret['变量1'])  # 修改页眉
            doc.replace_doc('变量1', ret['变量1'])
        elif i == 20:
            if ret['变量20']:
                doc.replace_doc('变量20', f"本项目免费维护期从：{ret['变量20']}开始，到{ret['变量21']}结束。")
            else:
                doc.replace_doc('变量20', '无')
        elif i == 22:
            doc.replace_doc('变量22', str(days(ret['变量11'], ret['变量6'])))
        elif i not in (20, 21, 22):
            doc.replace_doc('变量' + str(i), ret['变量' + str(i)])  # 替换其他变量


def docx_processing(file, path_prefix):
    if not os.path.exists(path_prefix + '\\处理完成'):
        os.makedirs(path_prefix + '\\处理完成')
    ret, ret2 = read_from_xlsx(file)
    # comm_task即可处理的文档：
    com_list = ['WN-QR-0-4-A 项目实施进度表-1.5.docx', 'WN-QR-2-12-A项目上线评估报告-1.5.docx', 'WN-QR-1-1-A 项目启动告客户书-1.5.docx',
                'WN-QR-0-1-A项目启动会会议记录-1.5.docx', 'WN-QR-1-5-A项目组成员清单-1.5.docx', '-----WN-QR-2-5-A数据准备与验收清单-1.5.docx']
    # comm_task之外需要另行做处理的文档：
    spc_list = ['WN-QR-4-3-A项目验收报告-1.5.docx', 'WN-QR-3-2-A系统切换方案-1.5.docx', 'WN-QR-2-4-A培训考核记录-1.5.docx',
                'WN-QR-2-3-A培训签到表-1.5.docx', 'WN-QR-2-1-B培训计划-1.5.docx', 'WN-QR-1-4-A项目实施计划-1.5.docx',
                '-----WN-QR-1-3-A项目功能范围确认单-1.5.docx', 'WN-QR-2-7-A工作底稿-1.5.docx', 'WN-QR-0-3-A软件及升级包杀毒记录-1.5.docx']
    if ret:
        for f in com_list:
            doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
            comm_task(doc, ret)
            doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
            doc.close()
            doc = None
            gc.collect()
        for f in spc_list:
            if f == 'WN-QR-4-3-A项目验收报告-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.ins_table_row(4, len(ret2) - 1)  # 根据子系统数量计算需要增加的表格行数
                for i in range(len(ret2)):  # 系统模块情况表内容填充
                    doc.modify_tab2(4, i + 3, 1, ret2[i][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-3-2-A系统切换方案-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.ins_table_row(2, len(ret2) - 1)  # 根据子系统数量计算需要增加的表格行数
                for i in range(len(ret2)):  # 系统模块情况表内容填充
                    doc.modify_tab(2, i + 2, 1, f"第{ret['变量15']}期")
                    doc.modify_tab(2, i + 2, 2, ret2[i][0])
                    doc.modify_tab(2, i + 2, 3, ret2[i][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-2-4-A培训考核记录-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.replace_doc('子系统（第一个）', ret2[0][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-2-3-A培训签到表-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.replace_doc('子系统（第一个）', ret2[0][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-2-1-B培训计划-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.replace_doc('变量a9', ret['变量9'][:7])
                doc.replace_doc('系统第一个', ret2[0][0])
                doc.replace_doc('子系统（第一个）', ret2[0][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-1-4-A项目实施计划-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.ins_table_row(2, len(ret2) - 1)  # 根据子系统数量计算需要增加的表格行数
                for i in range(len(ret2)):  # 系统模块情况表内容填充
                    doc.modify_tab(2, i + 2, 1, f"第{ret['变量15']}期")
                    doc.modify_tab(2, i + 2, 2, ret2[i][0])
                    doc.modify_tab(2, i + 2, 3, ret2[i][1])
                    doc.modify_tab(2, i + 2, 4, ret2[i][2])
                doc.modify_tab(3, 1, 2, ret['变量16'])
                begin = datetime.datetime.strptime(ret['变量11'], "%Y.%m.%d")
                end = datetime.datetime.strptime(ret['变量12'], "%Y.%m.%d")
                est_end = begin + datetime.timedelta(days=7)
                if est_end >= end:
                    rel_end = ret['变量12']
                    doc.modify_tab2(4, 10, 4, f"{ret['变量11']}到{rel_end}")
                    doc.modify_tab2(4, 11, 4, ret['变量12'])
                else:
                    rel_end = est_end.strftime('%Y{y}%m{m}%d{d}').format(y='.', m='.', d='')
                    doc.modify_tab2(4, 10, 4, f"{ret['变量11']}到{rel_end}")
                    doc.modify_tab2(4, 11, 4, f"{rel_end}到{ret['变量12']}")
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == '-----WN-QR-1-3-A项目功能范围确认单-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                doc.replace_doc('变量1', ret['变量1'])
                doc.ins_table_row(1, len(ret2) - 1)  # 根据子系统数量计算需要增加的表格行数
                for i in range(len(ret2)):  # 系统模块情况表内容填充
                    doc.modify_tab(1, i + 2, 1, ret2[i][0])
                    doc.modify_tab(1, i + 2, 2, ret2[i][1])
                    doc.modify_tab(1, i + 2, 3, ret2[i][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-2-7-A工作底稿-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.replace_doc('子系统（第一个）', ret2[0][1])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
            elif f == 'WN-QR-0-3-A软件及升级包杀毒记录-1.5.docx':
                doc = RemoteWord(path_prefix + '\\' + '套表模板\\' + f)
                comm_task(doc, ret)
                doc.replace_doc('系统（第一个）', ret2[0][0])
                doc.save_as(path_prefix + '\\处理完成' + '\\' + f)
                doc.close()
                doc = None
                gc.collect()
        sg.PopupOK('处理完成。', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray'))
    else:
        sg.PopupOK('未能正确获取数据，请检查数据来源文件。', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray'))


layout = [
    [sg.Text('数据来源：', background_color='#A8CFDD', font=("Microsoft YaHei Light", 12))],
    [sg.Input(size=(50, 2), font=("Microsoft YaHei Light", 12)),
     sg.FileBrowse(button_text='选择文件', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray'))],
    [sg.T(' ' * 80, background_color='#A8CFDD'),
     sg.Submit(size=(10, 1), button_text='开始处理', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray')),
     sg.Cancel(size=(10, 1), button_text='退出', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray'))],
]

sg.ChangeLookAndFeel('TealMono')
window = sg.Window('项目套表处理工具  v1.2 beta', icon="logo.ico").Layout(layout)

while True:
    button, values = window.Read()
    if button == '退出':
        window.Close()
    elif button is None:
        break
    elif button == '开始处理' and values[0] != '':
        path_prefix = os.getcwd()
        try:
            docx_processing(values[0], path_prefix)
        except Exception as e:
            sg.PopupOK(f'出错了：{str(e)}', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray'))
    else:
        sg.PopupOK('没有选择文件！', font=("Microsoft YaHei Light", 12), button_color=('white', 'gray'))
window.Close()
